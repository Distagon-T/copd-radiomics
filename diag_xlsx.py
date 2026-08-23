#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""诊断 E:/asthma.xlsx 读取失败原因"""
import os
import time

E = "asthma.xlsx"
D = "Asthma.xlsx"

for p in (E, D):
    print(f"{p}: size={os.path.getsize(p):,} bytes")
    with open(p, "rb") as f:
        head = f.read(8)
    print(f"  文件头: {head!r}  ->  {'有效 zip/xlsx' if head[:2]==b'PK' else '不是 xlsx!'}")
print()

# openpyxl read_only 探测 E 盘文件
import openpyxl
t = time.time()
try:
    wb = openpyxl.load_workbook(E, read_only=True, data_only=True)
    print(f"openpyxl read_only 打开成功: {wb.sheetnames}, 耗时 {time.time()-t:.1f}s")
    for ws in wb.worksheets:
        print(f"  sheet {ws.title}: max_row={ws.max_row}, max_col={ws.max_column}")
    wb.close()
except MemoryError as e:
    print(f"MemoryError: {e}")
except Exception as e:
    print(f"openpyxl 打开失败: {type(e).__name__}: {e}")
print(f"总耗时 {time.time()-t:.1f}s")
