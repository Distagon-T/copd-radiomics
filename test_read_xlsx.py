#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""隔离测试 read_xlsx_columns 的每一步"""
import sys, time

def log(*a):
    print(*a)
    sys.stdout.flush()

t0 = time.time()
try:
    import openpyxl
    log(f"[1] import openpyxl ok {time.time()-t0:.1f}s")
    wb = openpyxl.load_workbook(r"D:\copd-radiomics\Asthma.xlsx", read_only=True, data_only=True)
    log(f"[2] workbook open ok {time.time()-t0:.1f}s")
    ws = wb.worksheets[0]
    log(f"[3] sheet ok: {ws.max_row}x{ws.max_column} {time.time()-t0:.1f}s")
    it = ws.iter_rows(values_only=True)
    header = [str(h) if h is not None else "" for h in next(it)]
    log(f"[4] header done, n={len(header)} {time.time()-t0:.1f}s")
    keep_idx = [i for i, h in enumerate(header)
                if h in ("患者id", "年龄 (岁)") or "诊断" in h or "主诉" in h
                or "肌钙蛋白" in h or "NT_ProBNP" in h or "BNP" in h or "单位" in h]
    log(f"[5] keep cols={len(keep_idx)} {time.time()-t0:.1f}s")
    rows = []
    cnt = 0
    for row in it:
        rows.append([row[i] if i < len(row) else None for i in keep_idx])
        cnt += 1
        if cnt % 1000 == 0:
            log(f"[6] {cnt} rows {time.time()-t0:.1f}s mem={__import__('os').getpid()}")
    log(f"[7] iter done: {cnt} rows {time.time()-t0:.1f}s")
    wb.close()
    import pandas as pd
    df = pd.DataFrame(rows, columns=[header[i] for i in keep_idx])
    log(f"[8] DataFrame ok {len(df)}x{len(df.columns)} {time.time()-t0:.1f}s")
    print(df["患者id"].head(3).tolist())
except BaseException as e:
    import traceback
    traceback.print_exc()
    log(f"[FAIL] {type(e).__name__}: {e} at {time.time()-t0:.1f}s")
